import os
import threading
from tkinter import (
    Tk,
    filedialog,
    messagebox,
    Label,
    Entry,
    Button,
    StringVar,
    OptionMenu,
    IntVar,
    Checkbutton,
)
from tkinter.ttk import Progressbar
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import subprocess


def multiplicar_grupos_validos(
    archivo_excel, nombre_hoja, letra_columna, tiene_header=False
):
    wb = load_workbook(archivo_excel)
    ws = wb[nombre_hoja]

    col = column_index_from_string(letra_columna)
    max_col = ws.max_column
    max_row = ws.max_row
    grupos = []
    current_row = 2 if tiene_header else 1

    while current_row <= max_row:
        valor = ws.cell(row=current_row, column=col).value
        if valor not in (None, ""):
            inicio_grupo = current_row
            while current_row <= max_row and ws.cell(
                row=current_row, column=col
            ).value not in (None, ""):
                current_row += 1
            fin_grupo = current_row - 1

            inicio_espacios = current_row
            while current_row <= max_row and ws.cell(
                row=current_row, column=col
            ).value in (None, ""):
                current_row += 1
            fin_espacios = current_row - 1

            espacios = fin_espacios - inicio_espacios + 1
            if espacios > 0:
                grupos.append(
                    (inicio_grupo, fin_grupo, inicio_espacios, fin_espacios, espacios)
                )
        else:
            current_row += 1

    # Procesar los grupos en orden inverso para no afectar las posiciones siguientes
    for inicio, fin_grupo, inicio_espacios, fin_espacios, espacios in reversed(grupos):
        tamaño_bloque = fin_espacios - inicio + 1
        ws.insert_rows(fin_espacios + 1, espacios * tamaño_bloque)

        for copia in range(espacios):
            for i in range(tamaño_bloque):
                origen_row = inicio + i
                destino_row = fin_espacios + 1 + copia * tamaño_bloque + i
                for c in range(1, max_col + 1):
                    origen_valor = ws.cell(row=origen_row, column=c).value
                    ws.cell(row=destino_row, column=c).value = origen_valor

                # Si en la fila original la columna evaluada estaba vacía, también dejarla vacía en la copia
                if ws.cell(row=origen_row, column=col).value in (None, ""):
                    ws.cell(row=destino_row, column=col).value = None

    base, ext = os.path.splitext(archivo_excel)
    nuevo_nombre = f"{base}_duplicado{ext}"
    wb.save(nuevo_nombre)
    return nuevo_nombre


# -------------------- INTERFAZ --------------------


def seleccionar_archivo():
    ruta = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if ruta:
        archivo_var.set(ruta)
        cargar_hojas(ruta)


def cargar_hojas(ruta):
    try:
        wb = load_workbook(ruta)
        opciones_hojas.set(wb.sheetnames[0])
        menu_hojas["menu"].delete(0, "end")
        for hoja in wb.sheetnames:
            menu_hojas["menu"].add_command(
                label=hoja, command=lambda h=hoja: opciones_hojas.set(h)
            )
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo cargar el archivo: {e}")


def ejecutar_en_hilo():
    archivo = archivo_var.get()
    hoja = opciones_hojas.get()
    columna = columna_var.get().strip().upper()
    header = bool(tiene_header_var.get())

    if not os.path.isfile(archivo):
        messagebox.showerror("Error", "Selecciona un archivo válido.")
        return
    if not columna.isalpha():
        messagebox.showerror("Error", "Ingresa una letra de columna válida (A-Z).")
        return

    btn_ejecutar.config(state="disabled")
    label_proceso.config(text="Procesando... Esto puede tardar algunos segundos.")

    def proceso():
        try:
            nuevo = multiplicar_grupos_validos(archivo, hoja, columna, header)
            root.after(0, lambda: mostrar_opcion_abrir_archivo(nuevo))
        except Exception as err:
            root.after(
                0,
                lambda err=err: messagebox.showerror(
                    "Error durante el proceso", str(err)
                ),
            )
        finally:
            root.after(0, lambda: btn_ejecutar.config(state="normal"))
            root.after(0, lambda: label_proceso.config(text="Proceso completado."))

    threading.Thread(target=proceso).start()


def mostrar_opcion_abrir_archivo(archivo_guardado):
    respuesta = messagebox.askyesno(
        "Archivo guardado",
        f"El proceso ha finalizado.\nArchivo guardado como:\n\n{archivo_guardado}\n\n¿Quieres abrirlo?",
    )
    if respuesta:
        abrir_archivo(archivo_guardado)


def abrir_archivo(archivo):
    try:
        if os.name == "nt":  # Si es Windows
            subprocess.Popen(["start", archivo], shell=True)
        else:  # Si es Mac o Linux
            subprocess.call(["open", archivo])
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo abrir el archivo: {e}")


# -------------------- UI --------------------

root = Tk()
root.title("Duplicador de Grupos en Excel")
root.geometry("500x300")

archivo_var = StringVar()
columna_var = StringVar(value="A")
opciones_hojas = StringVar()
tiene_header_var = IntVar(value=1)

Label(root, text="Archivo Excel:").pack(pady=5)
Button(root, text="Seleccionar archivo", command=seleccionar_archivo).pack()
Label(root, textvariable=archivo_var, wraplength=450).pack(pady=5)

Label(root, text="Seleccionar hoja:").pack()
menu_hojas = OptionMenu(root, opciones_hojas, "")
menu_hojas.pack()

Label(root, text="Columna a evaluar (ej: A, B, C):").pack(pady=5)
Entry(root, textvariable=columna_var).pack()

Checkbutton(root, text="Tiene encabezado", variable=tiene_header_var).pack(pady=5)

btn_ejecutar = Button(
    root, text="Ejecutar duplicación", command=ejecutar_en_hilo, bg="lightgreen"
)
btn_ejecutar.pack(pady=10)

label_proceso = Label(root, text="", fg="blue")
label_proceso.pack(pady=10)


root.mainloop()
